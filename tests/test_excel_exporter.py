"""Comprehensive tests for the Excel exporter module.

This module contains tests for the ExcelExporter class, including:
- Basic export functionality
- Formatting and styling
- Handling of different data types
- Performance with various data sizes
- Error handling
"""
import os
import sys
import tempfile
import pytest
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Any

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from outlook_extractor.export.excel_exporter import ExcelExporter
from outlook_extractor.export.constants import EXPORT_FIELDS_V1

# Re-export fixtures from conftest for better IDE support
from tests.conftest import (
    sample_emails, empty_emails, invalid_emails,
    excel_exporter, create_test_emails
)

class TestExcelExporter:
    """Comprehensive test suite for the ExcelExporter class."""
    
    # --- Test Data ---
    
    @pytest.fixture
    def special_chars_email(self):
        """Email with special characters that might cause encoding issues."""
        return {
            'id': 'special-1',
            'conversation_id': 'conv-special',
            'subject': 'Email with special chars: 日本語, русский, 中文, ελληνικά',
            'sender_name': 'Tést Sènder',
            'sender_email': 'test.sender@example.com',
            'to_recipients': 'recipient@example.com',
            'cc_recipients': '',
            'bcc_recipients': '',
            'received_time': datetime.now() - timedelta(days=1),
            'sent_time': datetime.now() - timedelta(days=2),
            'categories': 'Test;Unicode;Special',
            'importance': 'Normal',
            'sensitivity': 0,
            'has_attachments': False,
            'is_read': True,
            'is_flagged': False,
            'is_priority': False,
            'is_admin': False,
            'body': 'This email contains special characters:\n• Bullet points\n• And emojis 😊\n• And quotes: "Hello"',
            'html_body': '<p>This email contains special characters:<br>• Bullet points<br>• And emojis 😊<br>• And quotes: &quot;Hello&quot;</p>',
            'folder_path': 'Inbox/Test',
            'thread_id': 'thread-special',
            'thread_depth': 0,
            'size': 1024
        }
    
    # --- Basic Export Tests ---
    
    @pytest.mark.parametrize("email_count", [1, 5, 10])
    def test_export_various_sizes(self, email_count, tmp_path, excel_exporter):
        """Test export with different numbers of emails."""
        # Setup
        output_file = tmp_path / f"test_export_{email_count}.xlsx"
        emails = create_test_emails(email_count)
        
        # Test
        result = excel_exporter.export_emails(
            emails=emails,
            output_path=output_file,
            include_headers=True
        )
        
        # Verify
        assert result is True
        assert output_file.exists()
        assert output_file.stat().st_size > 0
        
        # Check file contents
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check number of rows (header + emails)
        assert ws.max_row == email_count + 1  # +1 for header
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = list(EXPORT_FIELDS_V1)
        assert headers == expected_headers
    
    def test_export_with_special_chars(self, special_chars_email, tmp_path, excel_exporter):
        """Test export with special characters in email content."""
        # Setup
        output_file = tmp_path / "special_chars.xlsx"
        
        # Test
        result = excel_exporter.export_emails(
            emails=[special_chars_email],
            output_path=output_file,
            include_headers=True
        )
        
        # Verify
        assert result is True
        assert output_file.exists()
        
        # Check file contents
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check that special characters are preserved
        subject_cell = ws.cell(row=2, column=headers.index('subject') + 1)
        assert special_chars_email['subject'] in str(subject_cell.value)
    
    # --- Formatting Tests ---
    
    def test_export_with_formatting(self, sample_emails, tmp_path, excel_exporter):
        """Test export with custom formatting."""
        # Setup
        output_file = tmp_path / "formatted.xlsx"
        
        # Test with custom formatting
        excel_exporter.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        result = excel_exporter.export_emails(
            emails=sample_emails,
            output_path=output_file,
            include_headers=True
        )
        
        # Verify
        assert result is True
        
        # Check formatting
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Check header row formatting
        header_row = ws[1]
        for cell in header_row:
            assert cell.font.bold is True
            assert cell.font.color.rgb == '00FFFFFF'  # White
            assert cell.fill.start_color.rgb == '004F81BD'  # Blue
            assert cell.alignment.horizontal == 'center'
    
    # --- Error Handling Tests ---
    
    def test_export_empty_list(self, empty_emails, tmp_path, excel_exporter):
        """Test export with empty email list."""
        output_file = tmp_path / "empty_export.xlsx"
        
        with pytest.raises(ValueError, match="No emails to export"):
            excel_exporter.export_emails(
                emails=empty_emails,
                output_path=output_file,
                include_headers=True
            )
        
        # File should not be created
        assert not output_file.exists()
    
    def test_export_invalid_emails(self, invalid_emails, tmp_path, excel_exporter):
        """Test export with invalid email data."""
        output_file = tmp_path / "invalid_export.xlsx"
        
        with pytest.raises((ValueError, TypeError, AttributeError)):
            excel_exporter.export_emails(
                emails=invalid_emails,
                output_path=output_file,
                include_headers=True
            )
        
        # File should not be created
        assert not output_file.exists()
    
    # --- Performance Tests ---
    
    @pytest.mark.parametrize("email_count", [10, 100, 1000])
    def test_export_performance(self, email_count, tmp_path, excel_exporter, benchmark):
        """Test export performance with varying numbers of emails."""
        # Setup - create test emails
        emails = create_test_emails(email_count)
        output_file = tmp_path / f"perf_test_{email_count}.xlsx"
        
        # Benchmark the export
        def _run_export():
            return excel_exporter.export_emails(
                emails=emails,
                output_path=output_file,
                include_headers=True
            )
        
        # Run benchmark
        result = benchmark(_run_export)
        
        # Verify
        assert result is True
        assert output_file.exists()
        assert output_file.stat().st_size > 0
    
    # --- Integration Tests ---
    
    def test_export_then_read(self, sample_emails, tmp_path, excel_exporter):
        """Test that exported Excel file can be read back."""
        # Setup
        output_file = tmp_path / "roundtrip_test.xlsx"
        
        # Export to Excel
        result = excel_exporter.export_emails(
            emails=sample_emails,
            output_path=output_file,
            include_headers=True
        )
        assert result is True
        
        # Read back using pandas
        try:
            df = pd.read_excel(output_file, engine='openpyxl')
            assert len(df) == len(sample_emails)
            
            # Check that all expected columns are present
            for field in EXPORT_FIELDS_V1:
                assert field in df.columns
                
        except Exception as e:
            pytest.fail(f"Failed to read exported Excel file: {e}")
    
    # --- Worksheet Naming Tests ---
    
    def test_export_with_custom_sheet_name(self, sample_emails, tmp_path, excel_exporter):
        """Test export with custom sheet name."""
        output_file = tmp_path / "custom_sheet.xlsx"
        sheet_name = "CustomSheetName"
        
        result = excel_exporter.export_emails(
            emails=sample_emails,
            output_path=output_file,
            sheet_name=sheet_name,
            include_headers=True
        )
        
        assert result is True
        
        # Check sheet name
        wb = openpyxl.load_workbook(output_file)
        assert sheet_name in wb.sheetnames
    
    # --- File Operation Tests ---
    
    def test_export_to_nonexistent_directory(self, sample_emails, tmp_path, excel_exporter):
        """Test export to a non-existent directory."""
        output_file = tmp_path / "nonexistent" / "test_export.xlsx"
        
        result = excel_exporter.export_emails(
            emails=sample_emails,
            output_path=output_file,
            include_headers=True
        )
        
        assert result is True
        assert output_file.exists()
        assert output_file.parent.exists()
    
    def test_export_with_invalid_path(self, sample_emails, excel_exporter):
        """Test export with invalid file path."""
        # Test with None path
        with pytest.raises(ValueError):
            excel_exporter.export_emails(
                emails=sample_emails,
                output_path=None,
                include_headers=True
            )
        
        # Test with directory path
        with tempfile.TemporaryDirectory() as temp_dir:
            with pytest.raises(IsADirectoryError):
                excel_exporter.export_emails(
                    emails=sample_emails,
                    output_path=Path(temp_dir),
                    include_headers=True
                )
    
    # --- Helper Method Tests ---
    
    def test_format_cell_value(self, excel_exporter):
        """Test the _format_cell_value method."""
        # Test with different field types
        assert excel_exporter._format_cell_value(True) == 'Yes'
        assert excel_exporter._format_cell_value(False) == 'No'
        assert excel_exporter._format_cell_value(None) == ''
        
        # Test with datetime
        now = datetime.now()
        assert excel_exporter._format_cell_value(now) == now
        
        # Test with list
        assert excel_exporter._format_cell_value(['a', 'b', 'c']) == 'a, b, c'
        
        # Test with special characters
        assert excel_exporter._format_cell_value('Test "quotes"') == 'Test "quotes"'


if __name__ == "__main__":
    pytest.main(["-v", "test_excel_exporter.py"])
