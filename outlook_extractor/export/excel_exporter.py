"""Excel export functionality for Outlook Extractor.

This module provides functionality to export email data to Excel (XLSX) format
with proper formatting and multiple worksheets.
"""
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple, Callable
from threading import Event
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .. import constants

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Handles the export of email data to Excel format.
    
    This exporter creates well-formatted Excel files with multiple worksheets
    and proper styling for better readability.
    """
    
    def __init__(self, config: Optional[dict] = None):
        """Initialize the Excel exporter.
        
        Args:
            config: Optional configuration dictionary
        """
        self.config = config or {}
        self.fields = constants.EXPORT_FIELDS_V1
        self._setup_styles()
        self._setup_field_formatters()
    
    def _setup_styles(self) -> None:
        """Initialize Excel styles."""
        self.styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'alignment': Alignment(vertical='top', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='none'),
                    bottom=Side(style='thin')
                )
            },
            'summary': {
                'font': Font(italic=True, bold=True)
            }
        }
    
    def _setup_field_formatters(self) -> None:
        """Set up field formatters for Excel export."""
        self._field_formatters = {}
        
        # Define default formatters for different field types
        type_formatters = {
            'string': str,
            'number': float,  # Use float for numbers in Excel
            'boolean': bool,
            'datetime': lambda x: x if hasattr(x, 'isoformat') else str(x),
            'list': lambda x: '; '.join(str(i) for i in x) if isinstance(x, list) else str(x or ''),
            'dict': lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, dict) else str(x or ''),
            'text': str
        }
        
        # Set up formatters for each field
        for field in constants.EXPORT_FIELDS_V1:
            field_id = field['id']
            field_type = field.get('type', 'string')
            self._field_formatters[field_id] = type_formatters.get(field_type, str)
    
    def export_emails(
        self,
        emails: List[Dict[str, Any]],
        output_path: str,
        include_headers: bool = True,
        progress_callback: Optional[Callable[[int, int], None]] = None,
        cancel_event: Optional[Event] = None,
        fields: Optional[List[str]] = None
    ) -> Tuple[bool, str]:
        """Export emails to an Excel file.
        
        Args:
            emails: List of email dictionaries to export
            output_path: Path to the output Excel file
            include_headers: Whether to include headers in the output
            progress_callback: Optional callback for progress updates
            cancel_event: Optional event to cancel the export
            fields: Optional list of fields to include in the export
            
        Returns:
            Tuple[bool, str]: (success, message) - success status and message
        """
        if not emails:
            msg = "No emails to export"
            logger.warning(msg)
            return False, msg
            
        if cancel_event and cancel_event.is_set():
            msg = "Export cancelled by user"
            logger.info(msg)
            return False, msg
        
        output_path = Path(output_path)
        temp_path = output_path.with_suffix('.tmp' + output_path.suffix)
        total_emails = len(emails)
        
        try:
            # Create workbook and worksheets
            wb = Workbook()
            
            # Get field names from EXPORT_FIELDS_V1 if not provided
            if not fields:
                fields = [field['id'] for field in constants.EXPORT_FIELDS_V1]
            else:
                # Ensure fields are unique
                fields = list(dict.fromkeys(fields))
            
            # Remove default sheet and create our own
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create main data worksheet
            ws_data = wb.create_sheet("Emails")
            
            # Add headers
            if include_headers:
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Write headers
                for col_num, field_id in enumerate(fields, 1):
                    # Find the field in EXPORT_FIELDS_V1 to get the display name
                    field_name = field_id.replace('_', ' ').title()
                    for field in constants.EXPORT_FIELDS_V1:
                        if field['id'] == field_id:
                            field_name = field.get('name', field_name)
                            break
                        
                    cell = ws_data.cell(row=1, column=col_num, value=field_name)
                    cell.font = header_font
                    cell.border = thin_border
                    cell.fill = header_fill
            
            # Initialize counters
            processed_emails = 0
            error_count = 0
            max_errors = 10  # Maximum number of errors before giving up
            
            # Process emails in batches
            batch_size = 100
            for i in range(0, len(emails), batch_size):
                batch = emails[i:i + batch_size]
                
                # Process each email in the batch
                for email in batch:
                    if cancel_event and cancel_event.is_set():
                        return False, "Export cancelled by user"
                    
                    try:
                        # Format the email row
                        row_data = {}
                        for field in fields:
                            # Handle nested fields (e.g., 'sender.email_address')
                            if '.' in field:
                                parts = field.split('.')
                                value = email
                                for part in parts:
                                    if isinstance(value, dict):
                                        value = value.get(part, '')
                                    else:
                                        value = ''
                                        break
                            else:
                                value = email.get(field, '')
                            
                            # Apply formatter if available
                            formatter = self._field_formatters.get(field, str)
                            try:
                                row_data[field] = formatter(value) if value is not None else ''
                            except Exception as e:
                                logger.warning(f"Error formatting field '{field}': {e}")
                                row_data[field] = str(value) if value is not None else ''
                        
                        # Add the row to the worksheet in the correct field order
                        row = [row_data.get(field, '') for field in fields]
                        ws_data.append(row)
                        
                        # Update progress
                        processed_emails += 1
                        if progress_callback:
                            progress_callback(processed_emails, total_emails)
                            
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error processing email: {e}", exc_info=True)
                        if error_count >= max_errors:
                            return False, f"Too many errors ({error_count}) during export. Last error: {e}"
            
            # Auto-size columns
            self._auto_size_columns(ws_data)
            
            # Add summary sheet
            self._add_summary_sheet(wb, emails)
            
            # Save the workbook
            wb.save(temp_path)
            
            # Rename temp file to final name (atomic operation)
            if temp_path.exists():
                if output_path.exists():
                    output_path.unlink()
                temp_path.rename(output_path)
            
            msg = f"Successfully exported {len(emails)} emails to {output_path}"
            logger.info(msg)
            return True, msg
            
        except Exception as e:
            error_msg = f"Error exporting to Excel: {str(e)}"
            logger.error(error_msg, exc_info=True)
            
            # Clean up temp file if it exists
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception as cleanup_error:
                    logger.error(f"Failed to clean up temp file: {cleanup_error}")
            
            return False, error_msg
    
    def _auto_size_columns(self, worksheet) -> None:
        """Auto-size columns in the worksheet."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Find the maximum length of content in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = min(50, max(10, adjusted_width))
    
    def _add_summary_sheet(self, workbook, emails: List[Dict]) -> None:
        """Add a summary worksheet with statistics."""
        if not emails:
            return
        
        ws_summary = workbook.create_sheet("Summary")
        
        # Add title
        title_cell = ws_summary.cell(row=1, column=1, value="Email Export Summary")
        title_cell.font = Font(size=16, bold=True)
        
        # Add summary statistics
        stats = {
            "Total Emails": len(emails),
            "Unique Senders": len({e.get('sender_email', '') for e in emails if e.get('sender_email')}),
            "Emails with Attachments": sum(1 for e in emails if e.get('has_attachments')),
            "Average Recipients": sum(
                len(str(e.get('to_recipients', '')).split(';')) + 
                len(str(e.get('cc_recipients', '')).split(';')) + 
                len(str(e.get('bcc_recipients', '')).split(';'))
                for e in emails
            ) / max(1, len(emails))
        }
        
        # Add statistics to worksheet
        for i, (label, value) in enumerate(stats.items(), start=3):
            ws_summary.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=i, column=2, value=value)
        
        # Auto-size columns
        self._auto_size_columns(ws_summary)
        
        # Add a simple chart (requires openpyxl.chart)
        try:
            from openpyxl.chart import BarChart, Reference
            
            # Create a simple bar chart of email counts by sender
            sender_counts = {}
            for email in emails:
                sender = email.get('sender_email', 'Unknown')
                sender_counts[sender] = sender_counts.get(sender, 0) + 1
            
            # Sort senders by count and take top 10
            top_senders = dict(sorted(sender_counts.items(), 
                                    key=lambda x: x[1], 
                                    reverse=True)[:10])
            
            # Add data to worksheet
            ws_summary.cell(row=1, column=4, value="Top Senders").font = Font(bold=True)
            for i, (sender, count) in enumerate(top_senders.items(), start=2):
                ws_summary.cell(row=i, column=4, value=sender)
                ws_summary.cell(row=i, column=5, value=count)
            
            # Create and add chart
            chart = BarChart()
            chart.title = "Top 10 Senders"
            chart.y_axis.title = 'Number of Emails'
            chart.x_axis.title = 'Sender'
            
            data = Reference(ws_summary, 
                           min_col=5, min_row=1, 
                           max_row=len(top_senders)+1, 
                           max_col=5)
            cats = Reference(ws_summary, 
                           min_col=4, min_row=2, 
                           max_row=len(top_senders)+1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            
            # Position the chart
            ws_summary.add_chart(chart, "G2")
            
        except ImportError:
            logger.warning("openpyxl.chart not available, skipping chart creation")
        except Exception as e:
            logger.error(f"Error creating chart: {e}", exc_info=True)
